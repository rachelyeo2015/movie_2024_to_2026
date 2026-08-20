import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


df_movies = pd.read_csv("movies_2024_2026.csv") # without URLs columns
df_movies.info()    # 916 rows and 18 columns

# Create a scoring system ================
## 50% from box_office, 30% from ratings, 20% from vote_count to ensure reliability of vote
df_movies_copy = df_movies.copy()
df_movies_copy['box_office_usd_norm'] = df_movies_copy['box_office_usd'] / max(df_movies_copy['box_office_usd'])
df_movies_copy['rating_norm'] = df_movies_copy['rating_out_of_10'] / 10
df_movies_copy['vote_count_norm'] = df_movies_copy['vote_count'] / max(df_movies_copy['vote_count'])
df_movies_copy['score_norm_sum'] = df_movies_copy['box_office_usd_norm']*0.5 + df_movies_copy['rating_norm']*0.3 + df_movies_copy['vote_count_norm']*0.2

sorted_best_yearly_movies = df_movies_copy.sort_values(by=['release_year', 'score_norm_sum'], ascending=[True, False]).reset_index(drop=True)
sorted_best_yearly_movies['movie_rank_in_year'] = sorted_best_yearly_movies.groupby('release_year').cumcount() + 1

# Drop columns for easier analysis ===========
columns_to_drop = ['production_company','content_rating', 'screenwriter', 'language', 'country', 'synopsis','box_office_usd_norm', 'rating_norm', 'vote_count_norm']
sorted_best_yearly_movies_filtered = sorted_best_yearly_movies.drop(columns=columns_to_drop)

# Export excel for openpyxl formatting ===========
sorted_best_yearly_movies_filtered.to_excel("sorted_best_movies_2024_2026.xlsx", index=False, engine="openpyxl")


# Fill the rows of the top 10 best movies in each year based on scoring =============
wb_movies = load_workbook("sorted_best_movies_2024_2026.xlsx")
ws_movies = wb_movies.active

# going horizontally --> to get the header
header = [cell.value for cell in ws_movies[1]]  # always get value by calling ws.cell(row_no, col_no).value
rank_col_index = header.index('movie_rank_in_year') + 1 # because openpyxl starts from 1

# iterate through the rows with rank <= 10 and then iterate through each column within these rows to be filled
for row in range(2, ws_movies.max_row + 1): # start from 2 to skip the header row
    rank_cell_value = ws_movies.cell(row=row, column=rank_col_index).value
    if rank_cell_value <=10:
        for col in range(1, ws_movies.max_column + 1):
            ws_movies.cell(row=row, column=col).fill = PatternFill(fill_type='solid', start_color='FFFF00') # yellow fill

wb_movies.save("sorted_best_movies_2024_2026.xlsx")